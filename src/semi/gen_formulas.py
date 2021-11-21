#!/usr/bin/env python3
"""
This script is based on special pairs of varieties and subvarieties, as given by semi.xlsx. 
It generates all formulas from the given spreadsheet 
The aim is to find a model in the variety but not in the subvariety.

First non (1,1) line: 229
Last line: 3642
"""

import os
import sys
from ast import literal_eval as make_tuple
from openpyxl import load_workbook


comment_lines = ['% The aim is to find a model in the variety but not in the subvariety\n',
                 '% sos is the variety, and goals represent the subvariety.\n']

sos_line = "\nformulas(sos).\n"
goal_line = "\nformulas(goals).\n"
end_line = "end_of_list.\n"
basic_str = '(x * y) * z = x * (y * z).\n0*0=0.\n'
    

def read_data(excel_file):
    wb = load_workbook(excel_file)
    sheet_bases = wb["bases"]
    sheet_imply = wb["imply"]
    bases = {make_tuple(row_cells[0].value): row_cells[1].value for row_cells in sheet_bases.iter_rows()}
    implies = [[make_tuple(row_cells[0].value), make_tuple(row_cells[2].value)] for row_cells in sheet_imply.iter_rows()]
    return (bases, implies)


def write_file(out_dir, line_no, variety, subvariety, variety_formula, subvariety_formula):
    """ writes out a mace4 input file. "bottom" level implies "top" level, so the "goal"
        is the "bottom" level clause so to find a model in the "bigger" algebra but not in
        the smaller algebra.
    Args:
        out_dir (str): output directory
        line_no (int): line number in the excel file
        variety (str):  tuple representing the variety
        subvariety (str):  tuple representing the subvariety
        variety_formula (str): Mace4 formula for the variety
        subvariety_formula (str): Mace4 formula for the subvariety
    """
    fn = os.path.join(out_dir, f"{format(line_no, '04d')}_{subvariety[0]}_{subvariety[1]}_implies_{variety[0]}_{variety[1]}.in")
    with (open(fn, "w")) as fp:
        fp.writelines(comment_lines)
        fp.write(sos_line)
        fp.writelines(basic_str)
        fp.write(f"\n{variety_formula}\n")
        fp.write(end_line)
        fp.write(goal_line)
        fp.write(f"{subvariety_formula}\n")
        fp.write(end_line)


def gen_mace4_files(excel_file, varieties, out_dir):
    """
    Args:
        excel_file (str): full path name of the excel file containing the varieties and subvarieties
        varieties(List[int]): list of rows (as in the spreadsheet) to write out
        out_dir (str): output directory
    """
    bases, implies = read_data(excel_file)
    for line_no in varieties:
        subvariety, variety = implies[line_no-1]
        write_file(out_dir, line_no, variety, subvariety, bases[variety], bases[subvariety])
    

if __name__ == "__main__":
    excel_file = sys.argv[1]
    n1 = int(sys.argv[2])
    n2 = int(sys.argv[3])
    if n1 < 1 or n1 > n2:
        print("<from excel row> must not be greater than <to excel row>.")
    else:
        if len(sys.argv) > 4:
            out_dir = sys.argv[4]
        else:
            out_dir = "."
        v = gen_mace4_files(excel_file, range(n1, n2+1), out_dir)
    
